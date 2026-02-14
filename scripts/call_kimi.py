#!/usr/bin/env python3
"""
Kimi AI (Moonshot) CLI Helper - BuilTPro Brain AI Integration

Reusable CLI tool for calling Kimi/Moonshot API from the command line.
Supports multiple modes: chat, vision, code generation, document analysis.

Usage:
    python scripts/call_kimi.py --mode chat --prompt "Your question here"
    python scripts/call_kimi.py --mode vision --image-url <url> --prompt "Analyze this"
    python scripts/call_kimi.py --mode code --prompt "Generate EVM calculator"
    python scripts/call_kimi.py --mode document --file report.pdf --prompt "Summarize"

Environment Variables Required:
    MOONSHOT_API_KEY: Kimi/Moonshot API key

Optional Environment Variables:
    KIMI_MODEL: Model to use (default: kimi-k2.5)
    KIMI_MAX_TOKENS: Max response tokens (default: 4096)
    KIMI_TEMPERATURE: Temperature (default: 0.7)
    KIMI_BASE_URL: API base URL (default: https://api.moonshot.ai/v1)
"""

from __future__ import annotations

import argparse
import base64
import json
import logging
import os
import sys
from pathlib import Path

import httpx

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------


class KimiConfigurationError(ValueError):
    """Raised when Kimi client is misconfigured (e.g. missing API key)."""


class KimiAPIError(RuntimeError):
    """Raised when a Kimi API call fails."""


# ---------------------------------------------------------------------------
# System prompts per mode
# ---------------------------------------------------------------------------

SYSTEM_PROMPTS = {
    "chat": (
        "You are an AI assistant for the BuilTPro Brain AI construction management "
        "platform. You help with construction project management, scheduling, cost "
        "estimation, safety compliance, and engineering analysis for large-scale "
        "infrastructure projects."
    ),
    "vision": (
        "You are a construction site image analyst. Analyze the provided image and "
        "give detailed observations about: construction progress, safety compliance, "
        "material identification, structural elements, equipment usage, and any "
        "potential issues. Be precise and technical in your analysis."
    ),
    "code": (
        "You are a code generator for the BuilTPro Brain AI platform. Generate clean, "
        "well-documented code for construction data analysis tasks. The platform uses "
        "Python 3.11+, FastAPI, SQLAlchemy, pandas, numpy, and scikit-learn. Common "
        "tasks include: EVM (Earned Value Management) calculations, QTO (Quantity "
        "Take-Off) processing, schedule analysis (SPI/CPI), cost forecasting with "
        "Monte Carlo simulation, and IFC/BIM data processing."
    ),
    "document": (
        "You are a construction document analyst. Analyze the provided document text "
        "and extract key information including: action items, deadlines, responsible "
        "parties, cost figures, schedule milestones, risk items, change orders, and "
        "compliance requirements. Be thorough and structured in your analysis."
    ),
}

# ---------------------------------------------------------------------------
# File text extraction
# ---------------------------------------------------------------------------


def _extract_text_from_file(file_path: str) -> str:
    """Extract text content from PDF, DOCX, XLSX, or plain text files."""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    suffix = path.suffix.lower()

    if suffix == ".pdf":
        try:
            import fitz  # PyMuPDF

            doc = fitz.open(str(path))
            text = "\n".join(page.get_text() for page in doc)
            doc.close()
            return text
        except ImportError:
            raise KimiConfigurationError(
                "PyMuPDF (fitz) is required for PDF extraction. "
                "Install with: pip install PyMuPDF"
            )

    if suffix == ".docx":
        try:
            import docx

            doc = docx.Document(str(path))
            return "\n".join(p.text for p in doc.paragraphs)
        except ImportError:
            raise KimiConfigurationError(
                "python-docx is required for DOCX extraction. "
                "Install with: pip install python-docx"
            )

    if suffix == ".xlsx":
        try:
            import openpyxl

            wb = openpyxl.load_workbook(str(path), read_only=True)
            lines = []
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"--- Sheet: {sheet} ---")
                for row in ws.iter_rows(values_only=True):
                    lines.append("\t".join(str(c) if c is not None else "" for c in row))
            wb.close()
            return "\n".join(lines)
        except ImportError:
            raise KimiConfigurationError(
                "openpyxl is required for XLSX extraction. "
                "Install with: pip install openpyxl"
            )

    # Plain text fallback
    return path.read_text(encoding="utf-8", errors="replace")


# ---------------------------------------------------------------------------
# Kimi Client
# ---------------------------------------------------------------------------


class KimiClient:
    """Synchronous HTTP client for Kimi/Moonshot API."""

    def __init__(
        self,
        *,
        api_key: str | None = None,
        model: str | None = None,
        base_url: str | None = None,
        max_tokens: int | None = None,
        temperature: float | None = None,
    ) -> None:
        self.api_key = api_key or os.getenv("MOONSHOT_API_KEY", "").strip()
        if not self.api_key:
            raise KimiConfigurationError(
                "MOONSHOT_API_KEY is not set. "
                "Set it as an environment variable or pass api_key directly."
            )

        self.model = model or os.getenv("KIMI_MODEL", "kimi-k2.5")
        self.base_url = (base_url or os.getenv("KIMI_BASE_URL", "https://api.moonshot.ai/v1")).rstrip("/")
        self.max_tokens = max_tokens or int(os.getenv("KIMI_MAX_TOKENS", "4096"))
        self.temperature = temperature if temperature is not None else float(os.getenv("KIMI_TEMPERATURE", "0.7"))

    def complete(
        self,
        messages: list[dict],
        *,
        model: str | None = None,
        max_tokens: int | None = None,
        temperature: float | None = None,
        thinking: bool = False,
    ) -> dict:
        """Send a chat completion request and return the full response dict."""
        payload: dict = {
            "model": model or self.model,
            "messages": messages,
            "max_tokens": max_tokens or self.max_tokens,
            "temperature": temperature if temperature is not None else self.temperature,
        }
        if thinking:
            payload["extra_body"] = {"thinking": {"type": "enabled"}}

        headers = {
            "Authorization": f"Bearer {self.api_key}",
            "Content-Type": "application/json",
        }

        try:
            response = httpx.post(
                f"{self.base_url}/chat/completions",
                headers=headers,
                json=payload,
                timeout=120.0,
            )
            response.raise_for_status()
        except httpx.HTTPStatusError as exc:
            raise KimiAPIError(
                f"Kimi API returned {exc.response.status_code}: {exc.response.text}"
            ) from exc
        except httpx.HTTPError as exc:
            raise KimiAPIError(f"Kimi API request failed: {exc}") from exc

        return response.json()

    def chat(self, prompt: str, *, system_prompt: str | None = None, thinking: bool = False) -> str:
        """Simple text chat. Returns the assistant's reply."""
        messages = []
        if system_prompt:
            messages.append({"role": "system", "content": system_prompt})
        messages.append({"role": "user", "content": prompt})
        data = self.complete(messages, thinking=thinking)
        return data["choices"][0]["message"]["content"]

    def vision_analyze(
        self,
        prompt: str,
        *,
        image_url: str | None = None,
        image_path: str | None = None,
        thinking: bool = False,
    ) -> str:
        """Analyze an image via URL or local file path."""
        if not image_url and not image_path:
            raise KimiConfigurationError("Provide either --image-url or --image-path for vision mode")

        if image_path and not image_url:
            path = Path(image_path)
            if not path.exists():
                raise FileNotFoundError(f"Image not found: {image_path}")
            suffix = path.suffix.lower().lstrip(".")
            mime_map = {"jpg": "jpeg", "jpeg": "jpeg", "png": "png", "webp": "webp", "gif": "gif"}
            mime = mime_map.get(suffix, "png")
            encoded = base64.b64encode(path.read_bytes()).decode("utf-8")
            image_url = f"data:image/{mime};base64,{encoded}"

        messages = [
            {"role": "system", "content": SYSTEM_PROMPTS["vision"]},
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": image_url}},
                ],
            },
        ]
        data = self.complete(messages, thinking=thinking)
        return data["choices"][0]["message"]["content"]

    def code_generate(self, prompt: str, *, language: str = "python", thinking: bool = False) -> str:
        """Generate code for construction data analysis."""
        full_prompt = f"Generate {language} code for the following task:\n\n{prompt}"
        return self.chat(full_prompt, system_prompt=SYSTEM_PROMPTS["code"], thinking=thinking)

    def document_analyze(self, prompt: str, *, file_path: str | None = None, thinking: bool = False) -> str:
        """Analyze a document file or text content."""
        content_parts = [prompt]
        model_override = None

        if file_path:
            text = _extract_text_from_file(file_path)
            content_parts.append(f"\n\n--- DOCUMENT CONTENT ---\n{text}")
            # Use 128k model for large documents
            if len(text) > 20000:
                model_override = "moonshot-v1-128k"

        messages = [
            {"role": "system", "content": SYSTEM_PROMPTS["document"]},
            {"role": "user", "content": "\n".join(content_parts)},
        ]
        data = self.complete(messages, model=model_override, thinking=thinking)
        return data["choices"][0]["message"]["content"]


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Kimi AI (Moonshot) CLI Helper for BuilTPro Brain AI",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=(
            "Examples:\n"
            "  %(prog)s --mode chat --prompt 'What are EVM metrics?'\n"
            "  %(prog)s --mode vision --image-url https://example.com/site.jpg --prompt 'Analyze progress'\n"
            "  %(prog)s --mode code --prompt 'Generate EVM calculator script'\n"
            "  %(prog)s --mode document --file contract.pdf --prompt 'Extract action items'\n"
        ),
    )
    parser.add_argument("--mode", required=True, choices=["chat", "vision", "code", "document"], help="Operation mode")
    parser.add_argument("--prompt", required=True, help="User prompt / task description")
    parser.add_argument("--image-url", help="Image URL for vision mode")
    parser.add_argument("--image-path", help="Local image file path for vision mode")
    parser.add_argument("--file", help="Document file path for document mode")
    parser.add_argument("--language", default="python", help="Target language for code mode (default: python)")
    parser.add_argument("--model", help="Override Kimi model (default: env or kimi-k2.5)")
    parser.add_argument("--max-tokens", type=int, help="Override max tokens (default: env or 4096)")
    parser.add_argument("--temperature", type=float, help="Override temperature (default: env or 0.7)")
    parser.add_argument("--output", choices=["text", "json"], default="text", help="Output format (default: text)")
    parser.add_argument("--thinking", action="store_true", help="Enable Kimi thinking/reasoning mode")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging")

    args = parser.parse_args()

    if args.verbose:
        logging.basicConfig(level=logging.DEBUG, format="%(levelname)s: %(message)s")
    else:
        logging.basicConfig(level=logging.WARNING, format="%(levelname)s: %(message)s")

    try:
        client = KimiClient(
            model=args.model,
            max_tokens=args.max_tokens,
            temperature=args.temperature,
        )
    except KimiConfigurationError as exc:
        print(f"Configuration error: {exc}", file=sys.stderr)
        return 1

    try:
        if args.mode == "chat":
            result = client.chat(args.prompt, thinking=args.thinking)
        elif args.mode == "vision":
            result = client.vision_analyze(
                args.prompt,
                image_url=args.image_url,
                image_path=args.image_path,
                thinking=args.thinking,
            )
        elif args.mode == "code":
            result = client.code_generate(args.prompt, language=args.language, thinking=args.thinking)
        elif args.mode == "document":
            result = client.document_analyze(args.prompt, file_path=args.file, thinking=args.thinking)
        else:
            print(f"Unknown mode: {args.mode}", file=sys.stderr)
            return 1
    except FileNotFoundError as exc:
        print(f"File error: {exc}", file=sys.stderr)
        return 3
    except KimiConfigurationError as exc:
        print(f"Configuration error: {exc}", file=sys.stderr)
        return 1
    except KimiAPIError as exc:
        print(f"API error: {exc}", file=sys.stderr)
        return 2

    if args.output == "json":
        print(json.dumps({"mode": args.mode, "result": result}, indent=2))
    else:
        print(result)

    return 0


if __name__ == "__main__":
    sys.exit(main())
